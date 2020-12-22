import datetime
import openpyxl

YOUBI = ("月", "火", "水", "木", "金", "土", "日")

input_date = input("作成する年月度を入力してください（例：202001）：")

# 当月の初日
month_current = datetime.datetime.strptime(input_date + "01", "%Y%m%d")
# 前月の末日
month_last = month_current - datetime.timedelta(days=1)

# 勤怠管理の開始日と終了日（20日締め）
date_start = month_last.replace(day=16)
date_end = month_current.replace(day=15)

# エクセルブックファイル（ひな形）
file_template = "勤怠管理表.xlsx"
wb = openpyxl.load_workbook(file_template)

# エクセルシート
ws = wb["Sheet1"]

# 年月度の入力
ws.cell(row=1, column=1).value = month_current.year
ws.cell(row=1, column=3).value = month_current.month

# 月日曜日の入力
for row_num in range(13, 44):
    if date_start <= date_end:
        # 月
        ws.cell(row=row_num, column=1).value = date_start.month
        # 日
        ws.cell(row=row_num, column=2).value = date_start.day
        # 曜日
        ws.cell(row=row_num, column=3).value = YOUBI[date_start.weekday()]
    date_start += datetime.timedelta(days=1)

# 新規ブックファイル保存
file_new = month_current.strftime('%Y%m') + '_' + file_template
wb.save(file_new)